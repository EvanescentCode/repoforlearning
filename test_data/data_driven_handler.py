from openpyxl import load_workbook
from Automatyzacja.utilities.util import Utils
import logging


class DDT:

    """
    Klasa odpowiadająca za czytanie danych z pliku Excel.
    Przyjmuje argumenty: filename, sheet_name, number_of_tests.
    Zwraca listy zawierające dane z pliku Excel.
    Zmienia także pierwszy element list z int na str
    """

    def __init__(self, filename, sheet_name, number_of_tests):
        self.filename = filename
        self.sheet_name = sheet_name
        self.num_tests = number_of_tests
        self.data = []
        self.wb = None
        self.ws = None
        self.used_data = []
        self.used_data_str = []
        self.log = Utils().custom_logger(log_level=logging.DEBUG)

    def workbook_handling(self):
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[self.sheet_name]

    def save_workbook(self):
        self.wb.save(self.filename)

    def read_data(self):
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            self.data.append(list(row))
        # print("Dane odczytane:", self.data)
        self.save_workbook()
        return self.data

    def data_processing(self):
        count = 0
        for row_idx, item in enumerate(self.data, start=2):
            if item[2] == 'aktywny' and item[3] == 'N/A' and count < self.num_tests:
                item[3] = 'PASS'
                self.used_data.append(item)
                count += 1
                self.update_excel_data(row_idx, item)

        # print("Zaktualizowane dane:", self.data)
        return self.used_data

    def update_excel_data(self, row_idx, item):
        for col_idx, value in enumerate(item, start=1):
            self.ws.cell(row=row_idx, column=col_idx, value=value)

    def process_handler(self):
        self.workbook_handling()
        self.read_data()
        used_data = self.data_processing()
        # Zmiana nip na str
        for item in used_data:
            item[0] = str(item[0])
            self.log.info('____________________START____________________')
            self.log.debug(item)
        self.save_workbook()
        return used_data


